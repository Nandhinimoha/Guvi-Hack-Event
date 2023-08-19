# Methods to read and write an Excel file

from openpyxl import load_workbook
class Healthcare:
    def __init__(self,file_name,sheet_name):
        self.file = file_name
        self.sheet = sheet_name
    #  count the rows in an Excel file
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row
    # count the columns of an Excel file
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    # read data from an Excel file

    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

        # write data inside an Excel file

    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)